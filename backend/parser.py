import pandas as pd
import numpy as _np
import io
from datetime import timedelta
from collections import defaultdict

# Column name aliases — maps possible Teletrax/Reuters labels to our internal keys
COL_ALIASES = {
    'channel': ['Channel: Name', 'channel_name', 'Channel Name'],
    'market': ['Market: Name', 'market_name', 'Market Name'],
    # 'Region: Name' in Teletrax = channel's home region (United States, Portugal,
    # International — confirmed against actual exports). 'Market: Name' = the
    # broadcast DMA (Paducah, Sydney standalone, etc.), which is finer-grained.
    'region': ['Region: Name', 'Region Name', 'Channel: Region Name', 'Channel: Region name', 'channelRegionName'],
    'utc_start': ['UTC detection start', 'hitUtcDetectionStart', 'detection_start_date_time_utc'],
    'local_start': ['Local detection start', 'hitLocalDetectionStart', 'detection_start_date_time_local'],
    'story_id': ['Story ID', 'itemid', 'Item ID'],
    'slug': ['Slug line', 'slug', 'Slug'],
    'headline': ['Headline', 'headline', 'Title', 'title'],
    'detection_duration': ['Detection duration', 'hitDetectionDuration', 'detection_length'],
    'actual_length': ['Actual detection length', 'hitActualDetectionLength'],
    'asset_length': ['Asset: Length', 'assetLength'],
    'activation_date': ['Asset: Activation date start (UTC)', 'assetActivationDateStartUtc'],
}


def _resolve_columns(df):
    """Map dataframe columns to internal keys using aliases."""
    col_map = {}
    df_cols_lower = {c.lower().strip(): c for c in df.columns}
    for key, aliases in COL_ALIASES.items():
        for alias in aliases:
            if alias in df.columns:
                col_map[key] = alias
                break
            if alias.lower().strip() in df_cols_lower:
                col_map[key] = df_cols_lower[alias.lower().strip()]
                break
    return col_map


def _td_to_seconds(val):
    """Convert a timedelta, time, or string hh:mm:ss to total seconds."""
    if pd.isnull(val):
        return 0
    if isinstance(val, timedelta):
        return val.total_seconds()
    if hasattr(val, 'hour'):  # datetime.time
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, str):
        try:
            parts = val.split(':')
            if len(parts) == 3:
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
        except Exception:
            pass
    return 0


def _seconds_to_hms(seconds):
    """Convert seconds to human-readable string like '2m 34s' or '1h 4m'."""
    seconds = int(seconds)
    if seconds < 60:
        return f"{seconds}s"
    elif seconds < 3600:
        m, s = divmod(seconds, 60)
        return f"{m}m {s}s"
    else:
        h, rem = divmod(seconds, 3600)
        m = rem // 60
        return f"{h}h {m}m"


def _country_from_slug(slug):
    """First hyphen-separated rubric token of the slug — usually a country
    (USA, IRAN, JAPAN), sometimes a region (MIDEAST, EU) or topic
    (SOCCER, BUSINESS). Returns the raw token so producers see what they
    expect from the editorial slug. Falls back to 'Other'."""
    if not slug:
        return 'Other'
    import re
    s = str(slug).upper()
    s = re.sub(r'^(ADVISORY|CORRECTION)[-\s]+', '', s).strip()
    s = re.sub(r'^\d+[A-Z]{0,4}\s+', '', s)
    tokens = [t.strip() for t in s.split('/')[0].split('-') if t.strip()]
    return tokens[0] if tokens else 'Other'


def _slug_family(slug):
    """Topic stem of the slug — everything before the '/'.
    USA-SCREWWORM/EGGS-EMBARGO → 'USA-SCREWWORM'. Same ADVISORY /
    CORRECTION / numeric-prefix stripping as _country_from_slug, but
    keeps the full hyphenated stem rather than just the first token."""
    if not slug:
        return 'Other'
    import re
    s = str(slug).upper()
    s = re.sub(r'^(ADVISORY|CORRECTION)[-\s]+', '', s).strip()
    s = re.sub(r'^\d+[A-Z]{0,4}\s+', '', s)
    stem = s.split('/')[0].strip().strip('-')
    return stem if stem else 'Other'


def _region_from_slug(slug):
    """Derive editorial regions from slug (e.g. USA-CHINA/TARIFFS → {Americas, Asia Pacific}).
    Returns a set — stories can belong to multiple regions."""
    if not slug:
        return {'Other'}
    import re
    s = str(slug).upper()
    # Strip simple ADVISORY / CORRECTION prefixes
    s = re.sub(r'^(ADVISORY|CORRECTION)[-\s]+', '', s).strip()
    # Strip numeric legacy prefixes (e.g. "3128WD GERMANY...")
    s = re.sub(r'^\d+[A-Z]{0,4}\s+', '', s)
    # Check every hyphen-separated token in the topic (before the /)
    tokens = [t.strip() for t in s.split('/')[0].split('-') if t.strip()]

    AMERICAS = {
        'USA', 'US', 'TRUMP', 'BIDEN', 'OBAMA', 'CALIFORNIA', 'TEXAS', 'FLORIDA',
        'MIAMI', 'SEATTLE', 'ARIZONA', 'OKLAHOMA', 'HAWAII', 'COLORADO', 'WASHINGTON',
        'CANADA', 'MEXICO', 'CUBA', 'BRAZIL', 'ARGENTINA', 'COLOMBIA', 'PERU', 'CHILE',
        'BOLIVIA', 'VENEZUELA', 'ECUADOR', 'HAITI', 'PANAMA', 'HONDURAS', 'GUATEMALA',
        'SAN FRANCISCO', 'NEW YORK', 'LATIN', 'UNITEDHEALTH', 'AMAZON',
    }
    ASIA_PACIFIC = {
        'CHINA', 'JAPAN', 'INDIA', 'AUSTRALIA', 'TAIWAN', 'SOUTHKOREA', 'NORTHKOREA',
        'INDONESIA', 'PHILIPPINES', 'VIETNAM', 'MALAYSIA', 'THAILAND', 'SINGAPORE',
        'PAKISTAN', 'BANGLADESH', 'MYANMAR', 'LAOS', 'CAMBODIA', 'SRILANKA',
        'NEPAL', 'MONGOLIA', 'KAZAKHSTAN', 'KYRGYZSTAN', 'TAJIKISTAN', 'NEWZEALAND',
        'HONGKONG', 'MACAU', 'BYD', 'SAMSUNG', 'AUTOSHOW', 'PEGATRON', 'VINFAST',
        'SHINSEGAE', 'HUAWEI', 'XIAOMI', 'APEC', 'ASEAN', 'SOUTHCHINASEA', 'PANASIAN',
        'PAN ASIA', 'ASIA',
    }
    MIDDLE_EAST = {
        'IRAN', 'IRAQ', 'ISRAEL', 'ISRAELI', 'SAUDI', 'EGYPT', 'SYRIA', 'LEBANON',
        'JORDAN', 'KUWAIT', 'QATAR', 'UAE', 'EMIRATES', 'BAHRAIN', 'OMAN', 'YEMEN',
        'PALESTINE', 'PALESTINIANS', 'HAMAS', 'MIDEAST', 'GULF', 'NAHOST',
        'WARCRIMES', 'UNRWA', 'ISRAELPALESTINIANS', 'JERUSALEM',
        'TURKEY', 'AFGHAN', 'AFGHANISTAN',
    }
    AFRICA = {
        'AFRICA', 'SAFRICA', 'NIGERIA', 'KENYA', 'GHANA', 'ETHIOPIA', 'TANZANIA',
        'UGANDA', 'SENEGAL', 'MOROCCO', 'ALGERIA', 'TUNISIA', 'EGYPT',
        'CONGO', 'MALI', 'NIGER', 'SUDAN', 'SOMALIA', 'LIBYA', 'RWANDA',
        'ZIMBABWE', 'MOZAMBIQUE', 'MADAGASCAR', 'GAMBIA', 'IVORYCOAST',
        'WESTAFRICA', 'ANGOLA', 'CAMEROON', 'BURKINA', 'EBOLA',
    }
    EUROPE = {
        'RUSSIA', 'UKRAINE', 'GERMANY', 'FRANCE', 'BRITAIN', 'ITALY', 'SPAIN',
        'POLAND', 'SWEDEN', 'NORWAY', 'DENMARK', 'FINLAND', 'NETHERLANDS',
        'BELGIUM', 'AUSTRIA', 'SWITZERLAND', 'PORTUGAL', 'GREECE', 'HUNGARY',
        'ROMANIA', 'CZECH', 'BULGARIA', 'CROATIA', 'SERBIA', 'ALBANIA', 'KOSOVO',
        'LATVIA', 'LITHUANIA', 'ESTONIA', 'IRELAND', 'SCOTLAND', 'CYPRUS',
        'MALTA', 'LUXEMBOURG', 'SLOVAKIA', 'SLOVENIA', 'MONTENEGRO', 'MOLDOVA',
        'BELARUS', 'ARMENIA', 'AZERBAIJAN', 'GEORGIA', 'NORDIC', 'EU',
        'EUROZONE', 'ECB', 'NATO', 'EUROPE', 'EUROPEAN', 'DEUTSCHLAND',
        'BUNDESTAG', 'GROKO', 'FAESER', 'BUNDESWEHR', 'SPRITPREISE', 'TANKSTELLEN',
        'POPE', 'VATICAN', 'BALTIC', 'NAGORNO', 'MIGRATION',
    }

    regions = set()
    for token in tokens:
        if token in MIDDLE_EAST:
            regions.add('Middle East')
        elif token in AFRICA:
            regions.add('Africa')
        elif token in AMERICAS:
            regions.add('Americas')
        elif token in ASIA_PACIFIC:
            regions.add('Asia Pacific')
        elif token in EUROPE:
            regions.add('Europe')
    return regions if regions else {'Other'}


def parse_file(file_bytes, filename):
    """
    Parse a Teletrax CSV or XLSX export and return aggregated data.
    Returns dict with 'summary', 'stories', 'top_channels', 'top_markets', 'date_range'.
    """
    # Load into dataframe
    if filename.lower().endswith('.xlsx'):
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
    else:
        # Try common separators
        for sep in [',', '|', '\t']:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), sep=sep, skiprows=_detect_header_rows(file_bytes))
                if len(df.columns) > 3:
                    break
            except Exception:
                continue

    col = _resolve_columns(df)

    # Require at minimum a slug/story column and a channel column
    if 'slug' not in col and 'story_id' not in col:
        raise ValueError("Could not find story/slug columns in this file. Please check it is a Teletrax export.")
    if 'channel' not in col:
        raise ValueError("Could not find channel column in this file.")

    # Normalise key columns
    slug_col = col.get('slug') or col.get('story_id')
    df['_slug'] = df[slug_col].fillna('Unknown').astype(str).str.strip()
    df['_story_id'] = df[col['story_id']].fillna('').astype(str).str.strip() if 'story_id' in col else ''
    df['_headline'] = df[col['headline']].fillna('').astype(str).str.strip() if 'headline' in col else ''
    df['_channel'] = df[col['channel']].fillna('Unknown').astype(str).str.strip()
    df['_market'] = df[col['market']].fillna('Unknown').astype(str).str.strip() if 'market' in col else 'Unknown'
    # Broadcast-destination country (used for per-story "countries reached" count and
    # for the country-mix breakdown). Falls back to _market if region is missing for
    # an individual airing — reach-counting tolerates some noise.
    if 'region' in col:
        df['_country'] = df[col['region']].fillna('').astype(str).str.strip()
        df['_country'] = df['_country'].where(df['_country'] != '', df['_market'])
    else:
        df['_country'] = df['_market']

    # Channel's HOME country — must come strictly from Channel: Region Name. Falling
    # back to _market would mislabel channels (e.g. "France 24 English" detected in
    # an Italian DMA → labelled Italy). Empty string when truly missing so callers
    # render "Unknown" rather than guess.
    if 'region' in col:
        df['_channel_country'] = df[col['region']].fillna('').astype(str).str.strip()
    else:
        df['_channel_country'] = ''
    # Story-origin region derived from slug, not from broadcast region column

    # Parse time columns
    df['_actual_secs'] = df[col['actual_length']].apply(_td_to_seconds) if 'actual_length' in col else 0
    df['_detect_secs'] = df[col['detection_duration']].apply(_td_to_seconds) if 'detection_duration' in col else 0
    df['_asset_secs'] = df[col['asset_length']].apply(_td_to_seconds) if 'asset_length' in col else 0

    # Parse UTC datetime
    if 'utc_start' in col:
        df['_utc_start'] = pd.to_datetime(df[col['utc_start']], errors='coerce', utc=True)
    else:
        df['_utc_start'] = pd.NaT

    if 'activation_date' in col:
        df['_activation'] = pd.to_datetime(df[col['activation_date']], errors='coerce', utc=True)
    else:
        df['_activation'] = pd.NaT

    if 'local_start' in col:
        df['_local_start'] = pd.to_datetime(df[col['local_start']], errors='coerce')
    else:
        df['_local_start'] = pd.NaT

    # Date range of the whole dataset
    valid_dates = df['_utc_start'].dropna()
    date_range = {
        'from': valid_dates.min().strftime('%d %b %Y %H:%M UTC') if len(valid_dates) else 'Unknown',
        'to': valid_dates.max().strftime('%d %b %Y %H:%M UTC') if len(valid_dates) else 'Unknown',
    }
    dataset_end = valid_dates.max() if len(valid_dates) else None

    # Trend bin edges shared across all stories so sparklines are comparable.
    # <24h → hourly bins (up to 24); ≥24h → daily bins capped at 14.
    trend_edges = []
    trend_labels = []
    trend_unit = 'day'
    if len(valid_dates):
        span_hours = (valid_dates.max() - valid_dates.min()).total_seconds() / 3600
        if span_hours < 24:
            trend_unit = 'hour'
            n_bins = max(1, min(24, int(round(span_hours)) or 1))
            trend_edges = pd.date_range(valid_dates.min().floor('h'), periods=n_bins + 1, freq='h', tz='UTC')
        else:
            n_bins = min(14, max(2, int(span_hours // 24) + 1))
            trend_edges = pd.date_range(valid_dates.min().floor('D'), periods=n_bins + 1, freq='D', tz='UTC')
        trend_labels = [e.strftime('%d %b' if trend_unit == 'day' else '%H:%M') for e in trend_edges[:-1]]

    # Channel → home country lookup, computed once and reused everywhere a
    # channel needs to be labelled. Mode of non-empty Channel: Region Name,
    # "Unknown" when truly missing. Avoids the row-order-dependent garbage
    # the .iloc[0]/'first' approach produced.
    def _channel_home_country(s):
        non_empty = s[s != '']
        return non_empty.mode().iloc[0] if len(non_empty) else 'Unknown'
    channel_country_map = df.groupby('_channel')['_channel_country'].apply(_channel_home_country).to_dict()

    # --- Aggregate per story ---
    stories = []
    grouped = df.groupby('_slug', sort=False)

    for slug, grp in grouped:
        story_id = grp['_story_id'].iloc[0] if '_story_id' in grp.columns else ''
        headline_series = grp['_headline'][grp['_headline'] != ''] if '_headline' in grp.columns else pd.Series([], dtype=str)
        headline = headline_series.iloc[0] if len(headline_series) else ''
        airings = len(grp)
        channels = grp['_channel'].nunique()
        countries = grp['_country'].nunique()
        total_air_secs = grp['_actual_secs'].sum()
        avg_clip_secs = grp['_actual_secs'].mean()
        asset_secs = grp['_asset_secs'].iloc[0] if grp['_asset_secs'].iloc[0] > 0 else 0

        first_seen = grp['_utc_start'].min()
        last_seen = grp['_utc_start'].max()
        days_in_rotation = max(1, (last_seen - first_seen).days + 1) if pd.notna(first_seen) and pd.notna(last_seen) else 1

        # All channels for this story (used in modal channel breakdown).
        # Channel's home country comes from the precomputed map, not the row data.
        ch_agg = grp.groupby('_channel').agg(
            airings=('_actual_secs', 'count'),
            air_secs=('_actual_secs', 'sum'),
        ).reset_index().sort_values('airings', ascending=False)
        all_channels = ch_agg.to_dict('records')
        for c in all_channels:
            c['channel'] = c.pop('_channel')
            c['country'] = channel_country_map.get(c['channel'], 'Unknown')
            c['air_time'] = _seconds_to_hms(c['air_secs'])
            del c['air_secs']

        # All countries for this story (kept under all_markets for legacy compatibility)
        ctry_counts = grp.groupby('_country')['_actual_secs'].count().reset_index()
        ctry_counts.columns = ['market', 'airings']
        ctry_counts = ctry_counts.sort_values('airings', ascending=False)
        all_markets = ctry_counts.to_dict('records')

        # Regions derived from the story slug (not broadcast location); can be multiple
        regions = {r: 1 for r in _region_from_slug(slug)}

        # Trend sparkline: airings per shared time bucket
        if len(trend_edges):
            ts = grp['_utc_start'].dropna()
            if len(ts):
                counts, _ = _np.histogram(ts.astype('int64').to_numpy(),
                                          bins=trend_edges.astype('int64').to_numpy())
                trend = [int(c) for c in counts]
            else:
                trend = [0] * (len(trend_edges) - 1)
        else:
            trend = []

        # Longevity = % of airings OUTSIDE the first 24h after publish.
        # Higher = story had legs. Only meaningful if publish time is known
        # AND publish was ≥24h before the dataset's end, otherwise the
        # first-24h window is mechanically truncated and the figure misleads.
        longevity_pct = None
        publish_time = grp['_activation'].dropna().min() if '_activation' in grp.columns else pd.NaT
        airing_times = grp['_utc_start'].dropna()
        if pd.notna(publish_time) and len(airing_times) and dataset_end is not None:
            window_end = publish_time + pd.Timedelta(hours=24)
            if window_end <= dataset_end:
                first_24h = (airing_times <= window_end).sum()
                if airings > 0:
                    longevity_pct = round(100 * (1 - first_24h / airings))

        stories.append({
            'slug': slug,
            'headline': headline,
            'story_id': story_id,
            'airings': int(airings),
            'channels': int(channels),
            'countries': int(countries),
            'total_air_time': _seconds_to_hms(total_air_secs),
            'total_air_secs': int(total_air_secs),
            'avg_clip': _seconds_to_hms(avg_clip_secs),
            'avg_clip_secs': int(avg_clip_secs),
            'asset_length': _seconds_to_hms(asset_secs),
            'asset_secs': int(asset_secs),
            'first_seen': first_seen.strftime('%d %b %Y %H:%M') if pd.notna(first_seen) else '',
            'last_seen': last_seen.strftime('%d %b %Y %H:%M') if pd.notna(last_seen) else '',
            'days_in_rotation': int(days_in_rotation),
            'all_channels': all_channels,
            'all_markets': all_markets,
            'regions': regions,
            'trend': trend,
            'longevity': longevity_pct,
            'publish_time': publish_time.strftime('%d %b %Y %H:%M') if pd.notna(publish_time) else '',
            'publish_ts': int(publish_time.timestamp()) if pd.notna(publish_time) else None,
            'publish_day': publish_time.strftime('%Y-%m-%d') if pd.notna(publish_time) else None,
            'family': _slug_family(slug),
        })

    # Sort by airings descending
    stories.sort(key=lambda x: x['airings'], reverse=True)

    # --- Day context: what else was published on the same day? ---
    # Answers "did my story land on a crowded news day, or a quiet one?"
    # Groups by publish day (UTC) and slug family — IRAN-CRISIS, USA-TARIFFS, etc.
    # "The day's airings" deliberately means airings of stories *published* that
    # day, not all airings observed that day — we want the editorial weather at
    # publish, not every story still in rotation.
    day_context = {}
    if '_activation' in df.columns:
        pub = df.dropna(subset=['_activation']).copy()
        if len(pub):
            pub['_publish_day'] = pub['_activation'].dt.tz_convert('UTC').dt.strftime('%Y-%m-%d')
            pub['_family'] = pub['_slug'].map(_slug_family)
            fam_counts = pub.groupby(['_publish_day', '_family']).size().reset_index(name='airings')
            for day, day_grp in fam_counts.groupby('_publish_day'):
                day_grp = day_grp.sort_values('airings', ascending=False)
                total = int(day_grp['airings'].sum())
                top = [
                    {'family': r['_family'], 'airings': int(r['airings'])}
                    for _, r in day_grp.head(5).iterrows()
                ]
                day_context[day] = {
                    'total_airings': total,
                    'family_count': int(len(day_grp)),
                    'top_families': top,
                }

    # --- Aggregate per channel ---
    # One entry per channel: airings, distinct stories, air time, trend sparkline,
    # full list of stories aired (for the channel-detail modal).
    channels_out = []
    ch_grouped = df.groupby('_channel', sort=False)

    for chan, grp in ch_grouped:
        country = channel_country_map.get(chan, 'Unknown')
        ch_airings = len(grp)
        ch_stories = grp['_slug'].nunique()
        ch_air_secs = grp['_actual_secs'].sum()
        ch_avg_secs = grp['_actual_secs'].mean()

        first_seen = grp['_utc_start'].min()
        last_seen = grp['_utc_start'].max()
        days_active = max(1, (last_seen - first_seen).days + 1) if pd.notna(first_seen) and pd.notna(last_seen) else 1

        # Stories aired on this channel — one row per slug, sorted by airings desc
        s_agg = grp.groupby('_slug').agg(
            airings=('_actual_secs', 'count'),
            air_secs=('_actual_secs', 'sum'),
            story_id=('_story_id', 'first'),
            headline=('_headline', 'first'),
        ).reset_index().sort_values('airings', ascending=False)
        all_stories = s_agg.to_dict('records')
        for row in all_stories:
            row['slug'] = row.pop('_slug')
            row['air_time'] = _seconds_to_hms(row['air_secs'])
            del row['air_secs']
            # Pre-derive slug-based filtering fields so the client doesn't have
            # to mirror _region_from_slug/_country_from_slug. origin_country is
            # the exact token the pie aggregates by → click-to-filter matches 1:1.
            row['origin_country'] = _country_from_slug(row['slug'])
            row['regions'] = sorted(_region_from_slug(row['slug']))

        # Story-origin country mix for this channel (airings-weighted).
        # Pulled from the slug's first rubric token, so producers see the
        # same country labels they use editorially (USA, IRAN, JAPAN, etc).
        mix_counts = defaultdict(int)
        for row in all_stories:
            origin = row['origin_country']
            mix_counts[origin] += int(row['airings'])
        mix_sorted = sorted(mix_counts.items(), key=lambda kv: kv[1], reverse=True)
        # Top 8 + Other so the pie stays readable.
        if len(mix_sorted) > 8:
            top = mix_sorted[:8]
            other_total = sum(v for _, v in mix_sorted[8:])
            mix_sorted = top + [('Other', other_total)]
        story_country_mix = [{'country': k, 'airings': v} for k, v in mix_sorted]

        # Channel trend uses the same shared bin edges as stories — comparable sparklines.
        if len(trend_edges):
            ts = grp['_utc_start'].dropna()
            if len(ts):
                counts, _ = _np.histogram(ts.astype('int64').to_numpy(),
                                          bins=trend_edges.astype('int64').to_numpy())
                ch_trend = [int(c) for c in counts]
            else:
                ch_trend = [0] * (len(trend_edges) - 1)
        else:
            ch_trend = []

        channels_out.append({
            'channel': chan,
            'country': country,
            'airings': int(ch_airings),
            'stories': int(ch_stories),
            'total_air_time': _seconds_to_hms(ch_air_secs),
            'total_air_secs': int(ch_air_secs),
            'avg_clip': _seconds_to_hms(ch_avg_secs),
            'avg_clip_secs': int(ch_avg_secs),
            'first_seen': first_seen.strftime('%d %b %Y %H:%M') if pd.notna(first_seen) else '',
            'last_seen': last_seen.strftime('%d %b %Y %H:%M') if pd.notna(last_seen) else '',
            'days_active': int(days_active),
            'all_stories': all_stories,
            'story_country_mix': story_country_mix,
            'trend': ch_trend,
        })

    channels_out.sort(key=lambda x: x['airings'], reverse=True)

    # --- Global summary ---
    total_airings = len(df)
    total_stories = len(stories)
    total_channels = df['_channel'].nunique()
    total_countries = df['_country'].nunique()
    total_air_secs = df['_actual_secs'].sum()

    # Top 10 channels globally
    top_channels_global = (
        df.groupby('_channel')['_actual_secs']
        .agg(airings='count', air_secs='sum')
        .reset_index()
        .sort_values('airings', ascending=False)
        .head(10)
    )
    top_channels_global['air_time'] = top_channels_global['air_secs'].apply(_seconds_to_hms)
    top_channels_global['country'] = top_channels_global['_channel'].map(channel_country_map)
    top_channels_list = top_channels_global[['_channel', 'airings', 'air_time', 'country']].rename(
        columns={'_channel': 'channel'}
    ).to_dict('records')

    # Top 10 countries globally (kept under top_markets for legacy compatibility)
    top_markets_global = (
        df.groupby('_country')['_actual_secs']
        .agg(airings='count', air_secs='sum')
        .reset_index()
        .sort_values('airings', ascending=False)
        .head(10)
    )
    top_markets_global['air_time'] = top_markets_global['air_secs'].apply(_seconds_to_hms)
    top_markets_list = top_markets_global[['_country', 'airings', 'air_time']].rename(
        columns={'_country': 'market'}
    ).to_dict('records')

    # Zero-airing stories (if any — only possible if "include all assets" was ticked)
    zero_airing = sum(1 for s in stories if s['airings'] == 0)

    summary = {
        'total_airings': int(total_airings),
        'total_stories': int(total_stories),
        'total_channels': int(total_channels),
        'total_countries': int(total_countries),
        'total_air_time': _seconds_to_hms(total_air_secs),
        'zero_airing_stories': int(zero_airing),
        'date_range': date_range,
    }

    return {
        'summary': summary,
        'stories': stories,
        'channels': channels_out,
        'top_channels': top_channels_list,
        'top_markets': top_markets_list,
        'date_range': date_range,
        'trend_labels': trend_labels,
        'trend_unit': trend_unit,
        'day_context': day_context,
    }


def _detect_header_rows(file_bytes):
    """Detect if CSV has a metadata header block (Teletrax adds 3 lines before column headers)."""
    try:
        text = file_bytes.decode('utf-8', errors='ignore')
        lines = text.split('\n')
        for i, line in enumerate(lines[:5]):
            if 'Report date' in line or 'Starting' in line or 'Ending' in line:
                continue
            # First line that looks like column headers
            if ',' in line or '|' in line or '\t' in line:
                return i
    except Exception:
        pass
    return 0


def _slug_with_prefix(story):
    """Reuters producer-style slug: first 4 digits of story_id + slug.
    Mirrors the client's displaySlug() so exports match what's on screen."""
    sid = str(story.get('story_id') or '')
    slug = story.get('slug') or ''
    if sid:
        import re
        m = re.match(r'^\d{1,4}', sid)
        if m:
            return f"{m.group(0)}-{slug}"
    return slug


def generate_top_export(kind, rows, title=None):
    """Top-N export — schema depends on `kind`. Used by the contextual
    "Export top 25" button: whatever's on screen, sorted and filtered,
    lands in the spreadsheet in the same order."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active

    if kind == 'channels':
        ws.title = title or "Top Channels"
        headers = [
            'Channel', 'Country', 'Airings', 'Stories Aired',
            'Total Air Time', 'Avg Clip Used', 'Days Active',
            'First Seen', 'Last Seen', 'Top Story',
        ]
        col_widths = [30, 18, 10, 14, 16, 14, 12, 20, 20, 40]

        def row_for(c):
            top_story = (c.get('all_stories') or [{}])[0]
            top_slug = ''
            if top_story.get('slug'):
                # Reconstruct prefixed slug from story_id + slug like the UI does
                top_slug = _slug_with_prefix(top_story)
                if top_story.get('airings'):
                    top_slug = f"{top_slug} ({top_story['airings']} airings)"
            return [
                c.get('channel', ''),
                c.get('country', ''),
                c.get('airings', 0),
                c.get('stories', 0),
                c.get('total_air_time', ''),
                c.get('avg_clip', ''),
                c.get('days_active', 0),
                c.get('first_seen', ''),
                c.get('last_seen', ''),
                top_slug,
            ]
    else:
        ws.title = title or "Top Stories"
        headers = [
            'Story Slug', 'Headline', 'Airings', 'Channels', 'Countries',
            'Total Air Time', 'Avg Clip Used', 'Longevity', 'Date Published',
            'Days in Rotation', 'First Aired', 'Top Channel', 'Top Country',
        ]
        col_widths = [30, 50, 10, 10, 12, 16, 14, 12, 20, 12, 20, 25, 20]

        def row_for(s):
            top_ch = ''
            if s.get('all_channels'):
                first = s['all_channels'][0]
                top_ch = first.get('channel', '')
            top_mkt = s['all_markets'][0]['market'] if s.get('all_markets') else ''
            longevity = s.get('longevity')
            longevity_str = f"{longevity}%" if longevity is not None else ''
            return [
                _slug_with_prefix(s),
                s.get('headline', ''),
                s.get('airings', 0),
                s.get('channels', 0),
                s.get('countries', 0),
                s.get('total_air_time', ''),
                s.get('avg_clip', ''),
                longevity_str,
                s.get('publish_time', ''),
                s.get('days_in_rotation', 0),
                s.get('first_seen', ''),
                top_ch,
                top_mkt,
            ]

    header_fill = PatternFill(start_color='123015', end_color='123015', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 20

    for row_idx, item in enumerate(rows, 2):
        for col_idx, value in enumerate(row_for(item), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if row_idx % 2 == 0:
            light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_export(stories):
    """Generate a clean summary XLSX for download."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usage Summary"

    headers = [
        'Story Slug', 'Story ID', 'Airings', 'Channels', 'Countries',
        'Total Air Time', 'Avg Clip Used', 'Original Length',
        'Days in Rotation', 'First Aired', 'Last Aired',
        'Top Channel', 'Top Country'
    ]

    # Header row styling
    header_fill = PatternFill(start_color='123015', end_color='123015', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, name='Calibri', size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, story in enumerate(stories, 2):
        top_ch = story['all_channels'][0].get('channel', '') if story['all_channels'] else ''
        top_mkt = story['all_markets'][0]['market'] if story['all_markets'] else ''

        row_data = [
            story['slug'],
            story['story_id'],
            story['airings'],
            story['channels'],
            story['countries'],
            story['total_air_time'],
            story['avg_clip'],
            story['asset_length'],
            story['days_in_rotation'],
            story['first_seen'],
            story['last_seen'],
            top_ch,
            top_mkt,
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        # Alternate row shading
        if row_idx % 2 == 0:
            light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill

    # Column widths
    col_widths = [40, 20, 10, 10, 10, 15, 15, 15, 12, 18, 18, 25, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
